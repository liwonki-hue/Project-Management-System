# JM_Mapping.xlsx 템플릿 생성 스크립트
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL  = PatternFill('solid', fgColor='1D4ED8')
EXAMPLE_FILL = PatternFill('solid', fgColor='EFF6FF')
thin   = Side(style='thin', color='D0D5DD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 시트 1: JM Mapping ──────────────────────────────────────
ws = wb.active
ws.title = 'JM Mapping'

headers    = ['Activity ID', 'JM System', 'JM Unit', 'Description', 'Active']
col_widths = [22, 14, 12, 45, 10]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font      = Font(bold=True, color='FFFFFF', size=11)
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 28

examples = [
    ('0CF3010W58A11', 'RW', 'B0', 'B0 Raw Water System Piping DI', 'Y'),
]
for ri, row in enumerate(examples, 2):
    for ci, val in enumerate(row, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.fill      = EXAMPLE_FILL
        cell.alignment = Alignment(
            horizontal='left' if ci in (1, 4) else 'center',
            vertical='center'
        )
        cell.border = border
        cell.font   = Font(size=10)
    ws.row_dimensions[ri].height = 20

for ri in range(len(examples) + 2, len(examples) + 32):
    for ci in range(1, 6):
        cell = ws.cell(row=ri, column=ci)
        cell.border    = border
        cell.alignment = Alignment(
            horizontal='left' if ci in (1, 4) else 'center',
            vertical='center'
        )
        cell.font = Font(size=10)
    ws.row_dimensions[ri].height = 20

ws.freeze_panes = 'A2'

# ── 시트 2: Reference ───────────────────────────────────────
ws2 = wb.create_sheet('Reference')
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 25

ref_headers = [('A', 'B', 'System Code', 'System Description'),
               ('C', 'D', 'Unit Code',   'Unit Description')]
for ca, cb, ha, hb in ref_headers:
    for col, val in ((ca, ha), (cb, hb)):
        cell = ws2[f'{col}1']
        cell.value     = val
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

systems = [
    ('AS',      'Air Supply'),
    ('ATM',     'Atmosphere / Vent'),
    ('CCW',     'Closed Cooling Water'),
    ('CD',      'Condensate Drain'),
    ('DW',      'Domestic Water'),
    ('FG',      'Fuel Gas'),
    ('FGH',     'Fuel Gas High Pressure'),
    ('FO',      'Fuel Oil'),
    ('FW',      'Feed Water'),
    ('GT MISC', 'Gas Turbine Miscellaneous'),
    ('HP',      'High Pressure Steam'),
    ('HW',      'Hot Water'),
    ('IA',      'Instrument Air'),
    ('LO',      'Lube Oil'),
    ('LP',      'Low Pressure Steam'),
    ('N2',      'Nitrogen'),
    ('PW',      'Process Water'),
    ('RW',      'Raw Water'),
    ('SA',      'Service Air'),
    ('SS',      'Service Steam'),
    ('ST MISC', 'Steam Turbine Miscellaneous'),
    ('SW',      'Sea Water'),
    ('WWT',     'Waste Water Treatment'),
]
for ri, (code, desc) in enumerate(systems, 2):
    for ci, val in enumerate((code, desc), 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.border    = border
        cell.font      = Font(size=10)
        cell.alignment = Alignment(horizontal='center' if ci == 1 else 'left', vertical='center')

units = [('B0', 'Block 0 (0CF)'), ('B1', 'Block 1 (1CF)'),
         ('B2', 'Block 2 (2CF)'), ('E',  'E Block')]
for ri, (code, desc) in enumerate(units, 2):
    for ci, val in enumerate((code, desc), 3):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.border    = border
        cell.font      = Font(size=10)
        cell.alignment = Alignment(horizontal='center' if ci == 3 else 'left', vertical='center')

wb.save('JM_Mapping.xlsx')
print('JM_Mapping.xlsx 생성 완료')
