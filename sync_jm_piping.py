# JM(Joint Master) → PMS 피팅 데이터 동기화 스크립트
import os
import requests
import openpyxl
from datetime import date, timedelta
from dotenv import load_dotenv

load_dotenv()

SUPA_URL    = os.getenv("SUPABASE_URL")
SUPA_KEY    = os.getenv("SUPABASE_KEY")
BASE        = f"{SUPA_URL}/rest/v1"
MAPPING_XLS = "Raw File/JM_Mapping.xlsx"

PMS_HEADERS = {
    "apikey": SUPA_KEY,
    "Authorization": f"Bearer {SUPA_KEY}",
    "Accept-Profile":  "pms",
    "Content-Profile": "pms",
}
JM_HEADERS = {
    "apikey": SUPA_KEY,
    "Authorization": f"Bearer {SUPA_KEY}",
    "Accept-Profile": "construction",
}


def load_mapping() -> list:
    """JM_Mapping.xlsx 'JM Mapping' 시트에서 헤더명 기준으로 Active='Y' 행 읽기"""
    wb = openpyxl.load_workbook(MAPPING_XLS, data_only=True)
    ws = wb['JM Mapping']

    # 헤더 행에서 컬럼 인덱스 확인
    headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def col(name):
        for n in [name, name.replace(' ', '_'), name.replace(' ', '')]:
            try: return headers.index(n)
            except ValueError: pass
        return None

    i_id     = col('activity id')
    i_sys    = col('jm system')
    i_unit   = col('jm unit')
    i_ratio  = col('ratio')
    i_active = col('active')

    mapping = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        activity_id = str(row[i_id]   or '').strip() if i_id     is not None else ''
        system      = str(row[i_sys]  or '').strip() if i_sys    is not None else ''
        unit        = str(row[i_unit] or '').strip() if i_unit   is not None else ''
        active      = str(row[i_active] or '').strip().upper() if i_active is not None else ''

        if not activity_id or not system or not unit:
            continue
        if active != 'Y':
            continue

        ratio_val = row[i_ratio] if i_ratio is not None else None
        try:
            ratio = float(ratio_val) if ratio_val not in (None, '') else 1.0
        except (ValueError, TypeError):
            ratio = 1.0

        mapping.append((activity_id, system, unit, ratio))
    wb.close()
    return mapping


def week_bounds(d: date):
    """d를 포함하는 주의 (월요일, 일요일) 반환"""
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def fetch_jm(system: str, unit: str) -> list:
    """joint_master에서 해당 system+unit 전체 데이터 조회 (복수 system/unit 모두 + 구분자 지원)"""
    systems = [s.strip() for s in system.strip('()').split('+')]
    units   = [u.strip() for u in unit.strip('()').split('+')]
    params = {
        "select": "di,date_completed",
        "limit":  "5000",
    }
    params["system"] = f"eq.{systems[0]}" if len(systems) == 1 else f"in.({','.join(systems)})"
    params["unit"]   = f"eq.{units[0]}"   if len(units)   == 1 else f"in.({','.join(units)})"

    r = requests.get(f"{BASE}/joint_master", headers=JM_HEADERS, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def upsert_pms(table: str, records: list, on_conflict: str):
    h = {**PMS_HEADERS, "Content-Type": "application/json",
         "Prefer": "resolution=merge-duplicates,return=representation"}
    r = requests.post(f"{BASE}/{table}", headers=h,
                      params={"on_conflict": on_conflict}, json=records, timeout=30)
    r.raise_for_status()
    return r.json()


def sync():
    mapping = load_mapping()
    if not mapping:
        print(f"[오류] {MAPPING_XLS}에 Active='Y' 항목이 없습니다.")
        return
    print(f"매핑 로드 : {len(mapping)}개 항목")

    report_date     = date.today()
    report_date_str = report_date.isoformat()

    this_mon, this_sun = week_bounds(report_date)
    prev_mon = this_mon - timedelta(weeks=1)
    prev_sun = this_mon - timedelta(days=1)

    print(f"PMS Report date  : {report_date_str}")
    print(f"This week        : {this_mon} ~ {this_sun}")
    print(f"Previous week    : {prev_mon} ~ {prev_sun}")
    print()

    for activity_id, jm_system, jm_unit, ratio in mapping:
        print(f"[{activity_id}]  system={jm_system}, unit={jm_unit}, ratio={ratio}")

        joints = fetch_jm(jm_system, jm_unit)
        print(f"  JM 조회 결과 : {len(joints)}행")

        raw_total = sum(float(j.get("di") or 0) for j in joints)

        completed = [j for j in joints if j.get("date_completed")]
        start_date = (
            min(j["date_completed"][:10] for j in completed)
            if completed else None
        )

        # prev: 지난주 말(prev_sun)까지 누적 완료 물량 (cumulative)
        # this: 이번주에 완료된 물량만 (this week only)
        raw_prev = sum(
            float(j.get("di") or 0) for j in completed
            if j["date_completed"][:10] <= prev_sun.isoformat()
        )
        raw_this = sum(
            float(j.get("di") or 0) for j in completed
            if this_mon.isoformat() <= j["date_completed"][:10] <= this_sun.isoformat()
        )

        total_di = raw_total * ratio
        prev_di  = raw_prev  * ratio
        this_di  = raw_this  * ratio

        print(f"  Total DI (JM): {raw_total:.1f}  x {ratio} = {total_di:.1f}")
        print(f"  Start date   : {start_date or '-'}")
        print(f"  Prev week DI (cumul): {raw_prev:.1f}  x {ratio} = {prev_di:.1f}")
        print(f"  This week DI : {raw_this:.1f}  x {ratio} = {this_di:.1f}")

        # activities.budgeted_units + unit_type 갱신 (JM 매핑 항목은 항상 DI)
        upsert_pms("activities",
                   [{"activity_id": activity_id, "budgeted_units": total_di, "unit_type": "DI"}],
                   on_conflict="activity_id")
        print(f"  activities.budgeted_units = {total_di:.1f} 저장 완료")

        # weekly_progress 갱신
        prog = {
            "activity_id":   activity_id,
            "report_date":   report_date_str,
            "prev_week_qty": prev_di,
            "this_week_qty": this_di,
        }
        if start_date:
            prog["actual_start"] = start_date

        upsert_pms("weekly_progress", [prog], on_conflict="activity_id,report_date")
        print(f"  weekly_progress 저장 완료")
        print()

    print("=== 동기화 완료 ===")


if __name__ == "__main__":
    sync()
