# JM(Joint Master) 데이터를 PMS 진행률로 동기화하는 공용 로직 — 웹 요청/CLI 스크립트 공용
import os
import requests
import openpyxl
from datetime import date, timedelta
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

load_dotenv()

SUPA_URL    = os.getenv("SUPABASE_URL")
SUPA_KEY    = os.getenv("SUPABASE_KEY")
BASE        = f"{SUPA_URL}/rest/v1"
MAPPING_XLS = "Raw File/JM_Mapping.xlsx"

PMS_HEADERS = {
    "apikey": SUPA_KEY,
    "Authorization": f"Bearer {SUPA_KEY}",
    "Accept-Profile":  "pms",
    "Content-Profile": "pms",
}
JM_HEADERS = {
    "apikey": SUPA_KEY,
    "Authorization": f"Bearer {SUPA_KEY}",
    "Accept-Profile": "construction",
}


def load_mapping() -> list:
    """JM_Mapping.xlsx 'JM Mapping' 시트에서 헤더명 기준으로 Active='Y' 행 읽기"""
    wb = openpyxl.load_workbook(MAPPING_XLS, data_only=True)
    ws = wb['JM Mapping']

    headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def col(name):
        for n in [name, name.replace(' ', '_'), name.replace(' ', '')]:
            try: return headers.index(n)
            except ValueError: pass
        return None

    i_id, i_sys, i_unit, i_ratio, i_active = col('activity id'), col('jm system'), col('jm unit'), col('ratio'), col('active')

    mapping = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        activity_id = str(row[i_id]   or '').strip() if i_id     is not None else ''
        system      = str(row[i_sys]  or '').strip() if i_sys    is not None else ''
        unit        = str(row[i_unit] or '').strip() if i_unit   is not None else ''
        active      = str(row[i_active] or '').strip().upper() if i_active is not None else ''

        if not activity_id or not system or not unit:
            continue
        if active != 'Y':
            continue

        ratio_val = row[i_ratio] if i_ratio is not None else None
        try:
            ratio = float(ratio_val) if ratio_val not in (None, '') else 1.0
        except (ValueError, TypeError):
            ratio = 1.0

        mapping.append((activity_id, system, unit, ratio))
    wb.close()
    return mapping


def week_bounds_thu_wed(d: date):
    """d가 속한 목~수 주간의 (이번주 목요일, 이번주 수요일, 지난주 수요일) 반환.
    프론트엔드 getWeekWednesday()와 동일한 주 경계를 써야 report_date가 일치해
    주간 이월 로직(app.js buildProgressMap)과 충돌하지 않는다."""
    py_dow = (d.weekday() + 1) % 7  # JS Date.getDay()와 동일 (일=0 ... 토=6)
    offset = (3 - py_dow + 7) % 7
    this_wed = d + timedelta(days=offset)
    this_thu = this_wed - timedelta(days=6)
    prev_wed = this_thu - timedelta(days=1)
    return this_thu, this_wed, prev_wed


def fetch_jm(system: str, unit: str) -> list:
    """joint_master에서 해당 system+unit 전체 데이터 조회 (복수 system/unit 모두 + 구분자 지원)"""
    systems = [s.strip() for s in system.strip('()').split('+')]
    units   = [u.strip() for u in unit.strip('()').split('+')]
    params = {"select": "di,date_completed", "limit": "5000"}
    params["system"] = f"eq.{systems[0]}" if len(systems) == 1 else f"in.({','.join(systems)})"
    params["unit"]   = f"eq.{units[0]}"   if len(units)   == 1 else f"in.({','.join(units)})"

    r = requests.get(f"{BASE}/joint_master", headers=JM_HEADERS, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _build_records(entry, this_thu, this_wed, prev_wed, report_date_str):
    activity_id, jm_system, jm_unit, ratio = entry
    joints = fetch_jm(jm_system, jm_unit)

    raw_total = sum(float(j.get("di") or 0) for j in joints)
    completed = [j for j in joints if j.get("date_completed")]
    start_date = min((j["date_completed"][:10] for j in completed), default=None)

    raw_prev = sum(float(j.get("di") or 0) for j in completed if j["date_completed"][:10] <= prev_wed.isoformat())
    raw_this = sum(float(j.get("di") or 0) for j in completed
                   if this_thu.isoformat() <= j["date_completed"][:10] <= this_wed.isoformat())

    total_di, prev_di, this_di = raw_total * ratio, raw_prev * ratio, raw_this * ratio

    act_record = {"activity_id": activity_id, "budgeted_units": total_di, "unit_type": "DI"}
    # PostgREST 배치 upsert는 레코드마다 키 집합이 동일해야 하므로 actual_start도 항상 포함(없으면 null)
    prog_record = {
        "activity_id":   activity_id,
        "report_date":   report_date_str,
        "prev_week_qty": prev_di,
        "this_week_qty": this_di,
        "actual_start":  start_date,
    }

    return {
        "activity_id": activity_id, "total_di": total_di, "prev_di": prev_di,
        "this_di": this_di, "start_date": start_date,
        "act_record": act_record, "prog_record": prog_record,
    }


def _upsert(table: str, records: list, on_conflict: str):
    h = {**PMS_HEADERS, "Content-Type": "application/json",
         "Prefer": "resolution=merge-duplicates,return=minimal"}
    r = requests.post(f"{BASE}/{table}", headers=h,
                       params={"on_conflict": on_conflict}, json=records, timeout=30)
    r.raise_for_status()


def sync(max_workers: int = 10) -> dict:
    """활성 JM 매핑 전체를 동기화. 결과 레코드 리스트와 실패 목록을 반환한다."""
    mapping = load_mapping()
    if not mapping:
        return {"synced": 0, "failed": [], "results": []}

    this_thu, this_wed, prev_wed = week_bounds_thu_wed(date.today())
    report_date_str = this_wed.isoformat()

    results, failed = [], []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_build_records, entry, this_thu, this_wed, prev_wed, report_date_str): entry[0]
            for entry in mapping
        }
        for fut in futures:
            activity_id = futures[fut]
            try:
                results.append(fut.result())
            except Exception as e:
                failed.append(f"{activity_id}: {e}")

    if results:
        _upsert("activities", [r["act_record"] for r in results], on_conflict="activity_id")
        _upsert("weekly_progress", [r["prog_record"] for r in results], on_conflict="activity_id,report_date")

    return {"synced": len(results), "failed": failed, "results": results, "report_date": report_date_str}


def daily_breakdown_for_activity(activity_id: str) -> dict:
    """단일 활동의 이번 주(목~수) 일자별 완료 DI 합계 반환 (Piping 전용, 읽기 전용)"""
    mapping = load_mapping()
    entry = next((m for m in mapping if m[0] == activity_id), None)
    if not entry:
        return {}

    _, jm_system, jm_unit, ratio = entry
    this_thu, this_wed, _ = week_bounds_thu_wed(date.today())

    joints = fetch_jm(jm_system, jm_unit)
    breakdown = {}
    for j in joints:
        dc = j.get("date_completed")
        if not dc:
            continue
        d = dc[:10]
        if this_thu.isoformat() <= d <= this_wed.isoformat():
            breakdown[d] = breakdown.get(d, 0.0) + float(j.get("di") or 0) * ratio
    return breakdown
